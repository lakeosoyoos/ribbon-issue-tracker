"""
ribbon_parser.py
================
Reads EXFO-style reburn / splice report workbooks and extracts every flagged
event from the detailed ribbon x splice grid, classifying each by Legend type.

A "report" workbook has several sheets:
  - Acquisition Parameters
  - Reburn Summary       (per-splice / per-ribbon roll-up the tech already made)
  - Splice Report        (THE detailed ribbon x splice grid -- the source of truth)
  - Legend               (color -> meaning)
  - Distributed Loss     (optional)

This module finds the detailed grid sheet (handling the "more than one sheet"
case), parses each ribbon row and splice column, and returns a tidy list of
flagged events plus per-ribbon roll-ups so callers can aggregate across many
reports.

No Streamlit / pandas dependency here so it can be unit-tested or scripted.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

import openpyxl


# --------------------------------------------------------------------------- #
# Classification
# --------------------------------------------------------------------------- #
# Canonical issue categories (keys) -> human label.  Order is roughly
# severity / "needs action" descending.
CATEGORIES = {
    "reburn_ab":    "A+B Reburn (re-splice)",
    "break":        "Break / Broke",
    "ref":          "Reflective event (REF)",
    "launch":       "Launch / connector fault",
    "bend":         "Bend / damage",
    "gainer":       "Field gainer",
    "b_fill":       "B-fill (single-dir)",
    "a_only":       "A-only (single-dir)",
    "b_only":       "B-only (single-dir)",
    "dead_zone":    "Dead zone",
    "other":        "Other flagged",
}

# Last-6-hex fill color -> category, for cells whose text has no type token
# (a bare "fiber# .loss").  Derived from the Legend sheet.
COLOR_TO_CATEGORY = {
    "ffc7ce": "reburn_ab",   # Pink     - A+B bidirectional reburn
    "ffeb3b": "bend",        # Material yellow - BEND
    "ffff00": "bend",        # Pure yellow     - BEND
    "ffa500": "launch",      # Orange   - LAUNCH
    "ff9c27": "launch",
    "c6efce": "gainer",      # Mint green - field gainer
    "a9d08e": "gainer",
}

# Theme-indexed fills observed in the wild (theme, rounded-tint) -> category.
# theme 4 (accent1, blue) light tint  -> B-fill (light blue)
# theme 3 (lt2)           light tint  -> A-only / lavender / dead-zone family
THEME_TINT_TO_CATEGORY = {
    (4, 0.4): "b_fill",
    (4, 0.6): "b_fill",
    (3, 0.6): "a_only",
    (3, 0.4): "a_only",
}


def _color_key(fill) -> Optional[str]:
    """Return a normalized 6-char lowercase hex for a solid fill, or None."""
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    try:
        if fg.type == "rgb" and isinstance(fg.rgb, str):
            return fg.rgb[-6:].lower()
    except Exception:
        pass
    return None


def _theme_key(fill):
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    try:
        if fg.type == "theme":
            return (int(fg.theme), round(float(fg.tint or 0.0), 1))
    except Exception:
        pass
    return None


def classify_cell(value: str, fill) -> str:
    """Classify a flagged grid cell into a CATEGORIES key using text first,
    then fill color, then theme color."""
    t = str(value).lower()

    # --- text tokens (most reliable) ---
    if "bend" in t:
        return "bend"
    if "broke" in t or "break" in t:
        return "break"
    if "tailbox" in t or "launch" in t or "bad_" in t:
        return "launch"
    if "gainer" in t:
        return "gainer"
    if "dz" in t or "dead" in t:
        return "dead_zone"
    if "b-fill" in t or "(b-fill)" in t:
        return "b_fill"
    if re.search(r"\bref\b", t) or "(refl" in t:
        return "ref"
    if "(a)" in t:
        return "a_only"
    if "(b)" in t:
        return "b_only"
    if "bidi" in t:
        return "reburn_ab"

    # --- bare "fiber# .loss" cells: lean on color ---
    ck = _color_key(fill)
    if ck and ck in COLOR_TO_CATEGORY:
        return COLOR_TO_CATEGORY[ck]
    tk = _theme_key(fill)
    if tk and tk in THEME_TINT_TO_CATEGORY:
        return THEME_TINT_TO_CATEGORY[tk]

    # A pink-filled bare numeric is the classic averaged A+B reburn.
    if ck == "ffc7ce":
        return "reburn_ab"

    return "other"


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #
@dataclass
class Event:
    report: str
    route: str
    ribbon_num: Optional[int]
    tube: Optional[str]
    fiber_range: str
    splice: str
    category: str
    category_label: str
    text: str

    def as_row(self) -> dict:
        d = asdict(self)
        return d


@dataclass
class ReportParse:
    path: str
    report: str
    route: str
    grid_sheet: str
    sheets: list = field(default_factory=list)
    n_ribbons: int = 0
    n_splice_cols: int = 0
    events: list = field(default_factory=list)         # list[Event]
    summary_stats: dict = field(default_factory=dict)  # from Reburn Summary
    warnings: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #
_RIBBON_RE = re.compile(
    r"Fiber\s+(\d+)\s*-\s*(\d+)\s*\((\d+)\)\s*(?:\(([^)]+)\))?", re.IGNORECASE
)


def _parse_ribbon_label(label: str):
    """'Fiber 1-12 (1) (A1)' -> (range_str, ribbon_num, tube)."""
    if label is None:
        return None
    m = _RIBBON_RE.search(str(label))
    if not m:
        return None
    lo, hi, num, tube = m.groups()
    return (f"{lo}-{hi}", int(num), (tube or "").strip() or None)


def _find_grid_sheet(wb) -> Optional[str]:
    """Find the detailed ribbon x splice grid among possibly many sheets."""
    # 1) explicit name match
    for ws in wb.worksheets:
        if re.search(r"splice\s*report", ws.title, re.IGNORECASE):
            return ws.title
    # 2) structural match: a sheet whose col-A has a 'Ribbon' header followed
    #    by 'Fiber N-M (...)' rows.
    for ws in wb.worksheets:
        colA = [ws.cell(r, 1).value for r in range(1, min(ws.max_row, 12) + 1)]
        joined = " ".join(str(c) for c in colA if c)
        if "Ribbon" in joined and _RIBBON_RE.search(joined):
            return ws.title
    return None


def _find_header_row(ws) -> Optional[int]:
    for r in range(1, min(ws.max_row, 15) + 1):
        if str(ws.cell(r, 1).value).strip().lower() == "ribbon":
            return r
    return None


def _read_reburn_summary(wb) -> dict:
    """Pull the headline stats the tech's Reburn Summary already computed."""
    stats = {}
    for ws in wb.worksheets:
        if not re.search(r"reburn\s*summary", ws.title, re.IGNORECASE):
            continue
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            val = row[1] if len(row) > 1 else None
            if key in (
                "Ribbons", "Real splice columns", "Total ribbon × splice cells",
                "Cells with at least one reburn", "Reburn percentage",
            ):
                stats[key] = val
        break
    return stats


def parse_report(path: str | Path, report_name: str | None = None) -> ReportParse:
    """Parse one report workbook.

    report_name: optional display name. When the caller wrote `data` to a temp
    file (whose on-disk name is sanitized for the OS), pass the ORIGINAL file
    name here so the report/route labels reflect the real file — this is what
    lets us ingest a report regardless of how its filename is structured.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    report_name = (report_name or path.stem or "report").strip() or "report"
    route = _route_from_name(report_name)

    grid = _find_grid_sheet(wb)
    rp = ReportParse(
        path=str(path),
        report=report_name,
        route=route,
        grid_sheet=grid or "",
        sheets=list(wb.sheetnames),
    )
    rp.summary_stats = _read_reburn_summary(wb)

    if grid is None:
        rp.warnings.append("No ribbon x splice grid sheet found.")
        return rp

    ws = wb[grid]
    hdr = _find_header_row(ws)
    if hdr is None:
        rp.warnings.append(f"No 'Ribbon' header row in sheet '{grid}'.")
        return rp

    # Map data column -> splice label from the header row.
    splice_cols = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(hdr, c).value
        if v is not None and str(v).strip():
            splice_cols[c] = str(v).strip()
    rp.n_splice_cols = len(splice_cols)

    n_ribbons = 0
    for r in range(hdr + 1, ws.max_row + 1):
        parsed = _parse_ribbon_label(ws.cell(r, 1).value)
        if parsed is None:
            continue
        rng, num, tube = parsed
        n_ribbons += 1
        for c, splice_label in splice_cols.items():
            cell = ws.cell(r, c)
            val = cell.value
            if val is None or not str(val).strip():
                continue
            cat = classify_cell(val, cell.fill)
            rp.events.append(
                Event(
                    report=report_name,
                    route=route,
                    ribbon_num=num,
                    tube=tube,
                    fiber_range=rng,
                    splice=splice_label,
                    category=cat,
                    category_label=CATEGORIES.get(cat, cat),
                    text=str(val).strip(),
                )
            )
    rp.n_ribbons = n_ribbons
    return rp


def _route_from_name(name: str) -> str:
    """Best-effort human route name from the file stem."""
    s = re.sub(r"(?i)\b(reburn|oos|report|final|v?\d+(\.\d+)?|zk)\b", " ", name)
    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s or name


if __name__ == "__main__":
    import sys
    for p in sys.argv[1:]:
        rp = parse_report(p)
        print(f"\n{rp.report}  [route: {rp.route}]")
        print(f"  grid sheet: {rp.grid_sheet}  ribbons={rp.n_ribbons} "
              f"splice-cols={rp.n_splice_cols}  events={len(rp.events)}")
        for w in rp.warnings:
            print("  WARN:", w)
        for e in rp.events:
            print(f"    R{e.ribbon_num:>2} {e.tube or '--':<3} {e.splice:<16} "
                  f"{e.category_label:<22} {e.text}")
