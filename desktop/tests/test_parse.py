"""
test_parse.py — engine regression test (no customer data committed)

Builds a SYNTHETIC report workbook in a temp dir that mimics the EXFO-style
schema (Acquisition Parameters / Reburn Summary / Splice Report grid / Legend),
then asserts ribbon_parser:
  * finds the 'Splice Report' grid among the several sheets,
  * parses ribbon rows (Fiber N-M (k) (Tube)) + splice columns,
  * extracts flagged cells and classifies them by Legend type.

This is the CI gate for the parsing/classification engine. Runs before the
PyInstaller build so a parser regression fails fast.
"""
from __future__ import annotations

import os
import sys
import tempfile

import openpyxl
from openpyxl.styles import PatternFill

# import the engine from the repo root (one level up from desktop/tests)
_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(os.path.dirname(_HERE))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from ribbon_parser import parse_report  # noqa: E402

PINK = PatternFill("solid", fgColor="FFC7CE")   # A+B reburn (Legend swatch)
YELLOW = PatternFill("solid", fgColor="FFEB3B")  # bend (Legend swatch)
OFFPAL = PatternFill("solid", fgColor="9DB3D7")  # off-Legend tint (theme-ish)


def _make_synthetic(path: str) -> None:
    wb = openpyxl.Workbook()
    ap = wb.active
    ap.title = "Acquisition Parameters"
    ap["A1"] = "Parameter"; ap["B1"] = "Value"
    ap["A2"] = "Route"; ap["B2"] = "Synthetic Test Route"

    rs = wb.create_sheet("Reburn Summary")
    rs["A1"] = "Reburn Summary"; rs["B1"] = "3 ribbons × 2 splices"
    rs["A2"] = "Ribbons"; rs["B2"] = 3
    rs["A3"] = "Cells with at least one reburn"; rs["B3"] = 1

    sr = wb.create_sheet("Splice Report")
    # header row 3: col1 'Ribbon', then splice labels on odd columns
    sr.cell(3, 1, "Ribbon")
    sr.cell(3, 2, "ILA:A")
    sr.cell(3, 3, "Splice 1")
    sr.cell(3, 5, "Splice 2")
    sr.cell(3, 7, "ILA:B")
    # ribbon rows
    sr.cell(4, 1, "Fiber 1-12 (1) (A1)")
    sr.cell(5, 1, "Fiber 13-24 (2) (A2)")
    sr.cell(6, 1, "Fiber 25-36 (3) (B1)")
    # flagged cells
    c = sr.cell(4, 3, "5 .312")          # ribbon 1, Splice 1 — pink reburn
    c.fill = PINK
    c = sr.cell(6, 5, "30 BEND .112 bidi")  # ribbon 3, Splice 2 — bend
    c.fill = YELLOW
    # ribbon 2, Splice 1 — bare loss value with an off-Legend tint:
    # must be 'uncategorized', never guessed.
    c = sr.cell(5, 3, "13 .205")
    c.fill = OFFPAL

    lg = wb.create_sheet("Legend")
    lg["A1"] = "Color"; lg["B1"] = "Meaning"
    lg["A2"] = "Pink"; lg["B2"] = "A+B reburn"

    wb.save(path)


def test_parser_finds_grid_and_classifies():
    with tempfile.TemporaryDirectory() as tmp:
        p = os.path.join(tmp, "Synthetic Reburn Report.xlsx")
        _make_synthetic(p)
        rp = parse_report(p)

    assert rp.grid_sheet == "Splice Report", rp.grid_sheet
    assert rp.n_ribbons == 3, rp.n_ribbons
    assert len(rp.events) == 3, [e.text for e in rp.events]

    by_ribbon = {e.ribbon_num: e for e in rp.events}
    reburn = by_ribbon[1]
    assert reburn.tube == "A1"
    assert reburn.category == "reburn_ab", reburn.category   # exact Legend pink
    assert reburn.splice == "Splice 1"

    bend = by_ribbon[3]
    assert bend.tube == "B1"
    assert bend.category == "bend", bend.category            # text token

    # off-Legend tinted bare loss -> uncategorized, NOT a guessed type
    uncat = by_ribbon[2]
    assert uncat.tube == "A2"
    assert uncat.category == "uncategorized", uncat.category


if __name__ == "__main__":
    test_parser_finds_grid_and_classifies()
    print("test_parse: OK")
