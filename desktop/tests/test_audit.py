"""
test_audit.py — full click-through + ingestion audit.

Exercises every interactive path of ribbon_tracker_app.py through Streamlit's
AppTest harness (no browser) plus the ingestion edge cases, asserting the app
ingests a report regardless of how its filename is structured and never crashes
on bad input. Run: python desktop/tests/test_audit.py
"""
from __future__ import annotations

import io
import os
import sys
import glob
import zipfile
import tempfile

import openpyxl
from openpyxl.styles import PatternFill
from streamlit.testing.v1 import AppTest

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(os.path.dirname(_HERE))
for _p in (_ROOT, os.path.dirname(_HERE)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from ribbon_parser import parse_report  # noqa: E402

APP = os.path.join(_ROOT, "ribbon_tracker_app.py")
PINK, YELLOW = "FFC7CE", "FFEB3B"

_PASS, _FAIL = [], []


def check(name, cond, detail=""):
    (_PASS if cond else _FAIL).append(name)
    print(f"  [{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail and not cond else ""))


def make_report(path, cells, *, with_grid=True):
    """cells: list of (row, col, value, fill_hex)."""
    wb = openpyxl.Workbook()
    wb.active.title = "Acquisition Parameters"
    wb["Acquisition Parameters"]["A1"] = "Route"
    if with_grid:
        sr = wb.create_sheet("Splice Report")
        sr.cell(3, 1, "Ribbon")
        sr.cell(3, 3, "Splice 1")
        sr.cell(3, 5, "Splice 2")
        for r, lab in [(4, "Fiber 1-12 (1) (A1)"), (5, "Fiber 13-24 (2) (A2)"),
                       (6, "Fiber 25-36 (3) (B1)"), (7, "Fiber 37-48 (4) (B2)"),
                       (8, "Fiber 49-60 (5) (C1)"), (9, "Fiber 61-72 (6) (C2)")]:
            sr.cell(r, 1, lab)
        for r, c, v, f in cells:
            cell = sr.cell(r, c, v)
            if f:
                cell.fill = PatternFill("solid", fgColor=f)
    else:
        wb.create_sheet("Some Other Data")["A1"] = "nothing useful"
    wb.create_sheet("Legend")["A1"] = "Color"
    wb.save(path)


def run_app(folder=None, out_dir=None, autosave=None, deselect=None, twice=False):
    at = AppTest.from_file(APP, default_timeout=60)
    at.run()
    if folder is not None:
        at.session_state["folder_path"] = folder
    if out_dir is not None:
        for ti in at.text_input:
            if ti.label == "Save reports to":
                ti.set_value(out_dir)
    if autosave is not None:
        for cb in at.checkbox:
            if "Auto-save" in cb.label:
                cb.set_value(autosave)
    if deselect is not None:
        for ms in at.multiselect:
            if ms.label.startswith("Issue categories"):
                keep = [v for v in ms.value if v not in deselect]
                ms.set_value(keep)
    at.run()
    if twice:
        at.run()
    return at


# ─────────────────────────────────────────────────────────────────────────────
def audit_landing():
    print("\n[1] Landing page (no input)")
    at = AppTest.from_file(APP, default_timeout=60)
    at.run()
    check("landing renders, 0 exceptions", len(at.exception) == 0,
          str([e.value for e in at.exception]))
    check("shows 'add a report' prompt", any("Add at least one report" in i.value for i in at.info))


def audit_weird_filenames():
    print("\n[2] Weird filenames all ingest")
    with tempfile.TemporaryDirectory() as d:
        names = [
            "Normal Route.xlsx",
            "Réseau Nörð 1 (A→B).xlsx",          # unicode + arrow + parens
            "SEATTLE TO BEND.XLSX",               # uppercase extension
            "Span 5 reburn.xlsm",                 # .xlsm
            "Reburn Report Final v2.0.xlsx",      # only strip-tokens
            "12345.xlsx",                         # numeric only
            ("A_very_long_" + "x" * 120 + ".xlsx"),  # very long
        ]
        for i, nm in enumerate(names):
            row = 4 + (i % 6)
            make_report(os.path.join(d, nm), [(row, 3, f"{i+1} .3{i}0", PINK)])
        at = run_app(folder=d)
        labels = {m.label: m.value for m in at.metric}
        check("0 exceptions", len(at.exception) == 0, str([e.value for e in at.exception]))
        check(f"all {len(names)} weird-named reports loaded",
              labels.get("Reports loaded") == str(len(names)), str(labels))
        check("events parsed from them", int(labels.get("Total flagged events", "0")) >= len(names))


def audit_junk_skipped():
    print("\n[3] Junk files skipped (folder + zip)")
    with tempfile.TemporaryDirectory() as d:
        make_report(os.path.join(d, "Good.xlsx"), [(4, 3, "1 .310", PINK)])
        make_report(os.path.join(d, "~$Good.xlsx"), [(4, 3, "1 .310", PINK)])  # lock
        make_report(os.path.join(d, ".hidden.xlsx"), [(4, 3, "1 .310", PINK)])  # dotfile
        at = run_app(folder=d)
        labels = {m.label: m.value for m in at.metric}
        check("only the 1 real report loaded (lock/dotfile skipped)",
              labels.get("Reports loaded") == "1", str(labels))

    # zip with nested folders + __MACOSX junk.
    # Reproduce the app's extraction rule here (importing ribbon_tracker_app
    # would execute the whole Streamlit script outside its runtime).
    def _is_member(n):
        base = n.replace("\\", "/").rsplit("/", 1)[-1]
        return (n.lower().endswith((".xlsx", ".xlsm")) and not base.startswith("~$")
                and "__MACOSX" not in n and not base.startswith("."))

    def _xlsx_from_zip(zb):
        out = []
        with zipfile.ZipFile(io.BytesIO(zb)) as zf:
            for info in zf.infolist():
                if info.is_dir() or not _is_member(info.filename):
                    continue
                out.append((info.filename.rsplit("/", 1)[-1], zf.read(info)))
        return out

    buf = io.BytesIO()
    good = io.BytesIO()
    wb = openpyxl.Workbook(); wb.save(good)
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("Reports/A.xlsx", good.getvalue())
        zf.writestr("Reports/sub/B.xlsx", good.getvalue())
        zf.writestr("__MACOSX/Reports/._A.xlsx", b"junk")
        zf.writestr("Reports/.DS_Store", b"junk")
        zf.writestr("Reports/~$A.xlsx", b"junk")
    got = [n for n, _ in _xlsx_from_zip(buf.getvalue())]
    check("zip yields exactly the 2 real .xlsx (nested ok, junk dropped)",
          sorted(got) == ["A.xlsx", "B.xlsx"], str(got))


def audit_bad_input_no_crash():
    print("\n[4] No-grid + corrupt files never crash the app")
    with tempfile.TemporaryDirectory() as d:
        make_report(os.path.join(d, "Good.xlsx"), [(4, 3, "1 .310", PINK)])
        make_report(os.path.join(d, "No Grid.xlsx"), [], with_grid=False)
        with open(os.path.join(d, "Corrupt.xlsx"), "wb") as fh:
            fh.write(b"this is not a real xlsx file at all")
        at = run_app(folder=d)
        labels = {m.label: m.value for m in at.metric}
        check("app still renders, 0 unhandled exceptions",
              len(at.exception) == 0, str([e.value for e in at.exception]))
        check("the good report still loaded", int(labels.get("Reports loaded", "0")) >= 1, str(labels))
        warned = " ".join(w.value for w in at.warning)
        errored = " ".join(e.value for e in at.error)
        check("no-grid file surfaced a warning", "grid" in warned.lower(), warned)
        check("corrupt file surfaced an error (caught)", "Corrupt" in errored, errored)


def audit_duplicate_names():
    print("\n[5] Duplicate filenames handled")
    # same basename, DIFFERENT content, in root + subfolder -> both load
    with tempfile.TemporaryDirectory() as d:
        os.makedirs(os.path.join(d, "sub"))
        make_report(os.path.join(d, "Route.xlsx"), [(4, 3, "1 .310", PINK)])
        make_report(os.path.join(d, "sub", "Route.xlsx"), [(5, 5, "13 .288", PINK)])
        at = run_app(folder=d)
        labels = {m.label: m.value for m in at.metric}
        check("same-name different-content -> 2 reports",
              labels.get("Reports loaded") == "2", str(labels))
    # identical bytes twice -> deduped to 1
    with tempfile.TemporaryDirectory() as d:
        os.makedirs(os.path.join(d, "sub"))
        make_report(os.path.join(d, "Route.xlsx"), [(4, 3, "1 .310", PINK)])
        import shutil
        shutil.copy(os.path.join(d, "Route.xlsx"), os.path.join(d, "sub", "Route.xlsx"))
        at = run_app(folder=d)
        labels = {m.label: m.value for m in at.metric}
        check("identical file twice -> deduped to 1 report",
              labels.get("Reports loaded") == "1", str(labels))


def audit_tabs_and_filters():
    print("\n[6] Tabs render + category filter works")
    with tempfile.TemporaryDirectory() as d:
        make_report(os.path.join(d, "Alpha.xlsx"),
                    [(4, 3, "1 .310", PINK), (6, 5, "30 BEND .1 bidi", YELLOW)])
        make_report(os.path.join(d, "Beta.xlsx"),
                    [(4, 5, "2 .280", PINK)])
        at = run_app(folder=d)
        check("0 exceptions with full data", len(at.exception) == 0,
              str([e.value for e in at.exception]))
        check("repeat-offender callout present (Ribbon 1 in both)",
              any("Repeat offender" in i.value and "Ribbon 1" in i.value for i in at.info))
        check("4 tabs rendered", len(at.tabs) == 4, str(len(at.tabs)))
        full = {m.label: m.value for m in at.metric}
        # deselect bend -> fewer events
        at2 = run_app(folder=d, deselect=["bend"])
        filt = {m.label: m.value for m in at2.metric}
        check("deselecting 'bend' lowers total events",
              int(filt.get("Total flagged events", "0")) < int(full.get("Total flagged events", "0")),
              f"{full} vs {filt}")


def audit_autosave_and_clean():
    print("\n[7] Auto-save: once-per-analysis, clean route, toggle off")
    # normal -> writes one file, rerun does not duplicate
    with tempfile.TemporaryDirectory() as d, tempfile.TemporaryDirectory() as out:
        make_report(os.path.join(d, "Alpha.xlsx"), [(4, 3, "1 .310", PINK)])
        at = run_app(folder=d, out_dir=out, twice=True)
        files = glob.glob(os.path.join(out, "*.xlsx"))
        check("auto-saved exactly 1 file (dedup across reruns)", len(files) == 1, str(files))
        check("filename has date + time stamp",
              bool(files) and "Ribbon Issue Tracker 20" in os.path.basename(files[0]),
              str(files))
        if files:
            wb = openpyxl.load_workbook(files[0])
            check("report has the 3 expected sheets",
                  wb.sheetnames == ["Ribbon Ranking", "All Events", "Reports Loaded"],
                  str(wb.sheetnames))
    # clean route (no flagged events) STILL produces a saved report
    with tempfile.TemporaryDirectory() as d, tempfile.TemporaryDirectory() as out:
        make_report(os.path.join(d, "Clean.xlsx"), [])  # grid, zero flagged cells
        at = run_app(folder=d, out_dir=out)
        files = glob.glob(os.path.join(out, "*.xlsx"))
        check("clean route still writes a report", len(files) == 1, str(files))
    # autosave OFF -> no file
    with tempfile.TemporaryDirectory() as d, tempfile.TemporaryDirectory() as out:
        make_report(os.path.join(d, "Alpha.xlsx"), [(4, 3, "1 .310", PINK)])
        at = run_app(folder=d, out_dir=out, autosave=False)
        files = glob.glob(os.path.join(out, "*.xlsx"))
        check("auto-save OFF writes nothing", len(files) == 0, str(files))


def main():
    print("=" * 70)
    print("RIBBON TRACKER — full click-through + ingestion audit")
    print("=" * 70)
    audit_landing()
    audit_weird_filenames()
    audit_junk_skipped()
    audit_bad_input_no_crash()
    audit_duplicate_names()
    audit_tabs_and_filters()
    audit_autosave_and_clean()
    print("\n" + "=" * 70)
    print(f"RESULT: {len(_PASS)} passed, {len(_FAIL)} failed")
    if _FAIL:
        print("FAILED:")
        for f in _FAIL:
            print("  -", f)
        sys.exit(1)
    print("ALL CHECKS PASSED")


if __name__ == "__main__":
    main()
