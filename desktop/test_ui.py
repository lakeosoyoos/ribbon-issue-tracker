"""
test_ui.py — UI smoke test via Streamlit AppTest (no browser, ~10s).

Checks:
  * the launcher's PORT constant is still 8512 (port-registry guard),
  * ribbon_tracker_app.py renders its landing page with NO exceptions,
  * a synthetic report loaded through the folder-path input produces the
    expected metrics + repeat-offender plumbing (the cross-report roll-up).

Runs in CI BEFORE the PyInstaller build so a UI regression fails fast.
"""
from __future__ import annotations

import os
import sys
import tempfile

import openpyxl
from openpyxl.styles import PatternFill
from streamlit.testing.v1 import AppTest

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
for _p in (_ROOT, _HERE):
    if _p not in sys.path:
        sys.path.insert(0, _p)

APP = os.path.join(_ROOT, "ribbon_tracker_app.py")


def _check_port():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "rt_launcher", os.path.join(_HERE, "launcher.py"))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    assert mod.PORT == 8512, f"PORT drifted to {mod.PORT}; update the registry"
    print("port guard: PORT == 8512 OK")


def _make_report(path, route_tube_events):
    """route_tube_events: list of (row, ribbon_label, col, value, fill)."""
    wb = openpyxl.Workbook()
    wb.active.title = "Acquisition Parameters"
    sr = wb.create_sheet("Splice Report")
    sr.cell(3, 1, "Ribbon")
    sr.cell(3, 3, "Splice 1")
    sr.cell(3, 5, "Splice 2")
    for r, label in [(4, "Fiber 1-12 (1) (A1)"),
                     (5, "Fiber 13-24 (2) (A2)"),
                     (6, "Fiber 25-36 (3) (B1)")]:
        sr.cell(r, 1, label)
    for r, _lbl, c, val, fill in route_tube_events:
        cell = sr.cell(r, c, val)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
    wb.create_sheet("Legend")
    wb.save(path)


def _landing_renders():
    at = AppTest.from_file(APP, default_timeout=40)
    at.run()
    assert len(at.exception) == 0, [e.value for e in at.exception]
    print("landing page: 0 exceptions OK")


def _folder_rollup():
    with tempfile.TemporaryDirectory() as tmp:
        # two reports, ribbon 1 (A1) flagged in BOTH -> repeat offender
        _make_report(os.path.join(tmp, "Route Alpha.xlsx"),
                     [(4, "", 3, "5 .312", "FFC7CE")])
        _make_report(os.path.join(tmp, "Route Beta.xlsx"),
                     [(4, "", 5, "7 .280", "FFC7CE"),
                      (6, "", 3, "30 BEND .1 bidi", "FFEB3B")])
        at = AppTest.from_file(APP, default_timeout=40)
        at.run()
        at.session_state["folder_path"] = tmp
        at.run()
        assert len(at.exception) == 0, [e.value for e in at.exception]
        labels = {m.label: m.value for m in at.metric}
        assert labels.get("Reports loaded") == "2", labels
        assert labels.get("Total flagged events") == "3", labels
        infos = " ".join(i.value for i in at.info)
        assert "Repeat offender" in infos and "Ribbon 1" in infos, infos
        print("folder roll-up: 2 reports, repeat offender Ribbon 1 OK")


def main():
    _check_port()
    _landing_renders()
    _folder_rollup()
    print("test_ui: ALL OK")


if __name__ == "__main__":
    main()
