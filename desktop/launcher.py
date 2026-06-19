"""
Ribbon Tracker — PyInstaller launcher
=====================================
Entry point of the frozen Windows .exe / Mac .app. It:

  1. Redirects stdout/stderr to a log file (a frozen windowed app's stdout is
     None and any print() crashes the process).
  2. Pre-seeds Streamlit's first-run prompt + headless env vars so the hidden
     subprocess doesn't block on stdin.
  3. Optionally auto-updates the engine + UI files from raw.githubusercontent
     /main into ~/.ribbonTracker/engine, validated all-or-nothing (non-empty +
     contains b"def " + compile()). On any failure, falls back to the bundled
     copies that ship inside the build.
  4. Spawns a thread that waits until Streamlit answers /_stcore/health = "ok",
     then opens the browser.
  5. Detects an already-running instance and just opens a new tab into it.
  6. Boots Streamlit via its CLI entry point.

NOTE — the auto-update mechanism runs AFTER the bundle bootstraps, so it can
only ship changes to the .py files; it can never fix a bundle that won't boot.
Any change to launcher.py, the .spec, Python version, or bundled deps needs a
fresh build download from the GitHub Release. Auto-update also only works while
the repo (raw files) is publicly reachable; against a private repo every fetch
404s and the app runs the bundled engine — which is the safe default.
"""
from __future__ import annotations

import os
import sys
import time
import threading
import socket
import shutil
import urllib.request
import urllib.error
import webbrowser
from pathlib import Path


APP_NAME       = "RibbonTracker"
APP_DIR_NAME   = ".ribbonTracker"
HOST           = "127.0.0.1"
# Each desktop app must claim its OWN port so the launchers don't see each
# other's running server via /_stcore/health and silently open the wrong app.
# Registry: Secret Sauce 8501, SpliceReport 8503, Uni 8505, OTDR Suite 8510.
# Ribbon Tracker runs on 8512.
PORT           = 8512
HEALTH_URL     = f"http://{HOST}:{PORT}/_stcore/health"
APP_URL        = f"http://{HOST}:{PORT}"
GH_OWNER       = "lakeosoyoos"
GH_REPO        = "ribbon-issue-tracker"
GH_BRANCH      = "main"
RAW_URL_FMT    = ("https://raw.githubusercontent.com/"
                  f"{GH_OWNER}/{GH_REPO}/{GH_BRANCH}/{{path}}")

# Files we fetch from main on each launch. Order matters for the import smoke
# check: ribbon_parser must compile before the UI that imports it.
ENGINE_FILES = [
    "ribbon_parser.py",
    "error_reporter.py",
    "ribbon_tracker_app.py",
]


# ─────────────────────────────────────────────────────────────────────────────
#  1. Redirect stdout / stderr (frozen windowed exe has None for both)
# ─────────────────────────────────────────────────────────────────────────────
def _redirect_output_to_log() -> Path:
    log_dir = Path.home() / APP_DIR_NAME
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"{APP_NAME.lower()}.log"
    fh = open(log_path, "a", buffering=1, encoding="utf-8", errors="replace")
    sys.stdout = fh
    sys.stderr = fh
    print(f"\n=== {APP_NAME} launch {time.strftime('%Y-%m-%d %H:%M:%S')} ===")
    print(f"frozen={getattr(sys, 'frozen', False)}  executable={sys.executable}")
    return log_path


# ─────────────────────────────────────────────────────────────────────────────
#  2. Silence Streamlit's first-run email prompt (blocks on stdin)
# ─────────────────────────────────────────────────────────────────────────────
def _silence_first_run_prompt() -> None:
    cred_dir = Path.home() / ".streamlit"
    cred_dir.mkdir(parents=True, exist_ok=True)
    cred_path = cred_dir / "credentials.toml"
    if not cred_path.exists():
        cred_path.write_text('[general]\nemail = ""\n', encoding="utf-8")
    os.environ.setdefault("STREAMLIT_SERVER_HEADLESS", "true")
    os.environ.setdefault("STREAMLIT_BROWSER_GATHER_USAGE_STATS", "false")
    os.environ.setdefault("STREAMLIT_GLOBAL_DEVELOPMENT_MODE", "false")
    os.environ.setdefault("STREAMLIT_SERVER_ADDRESS", HOST)
    os.environ.setdefault("STREAMLIT_SERVER_PORT", str(PORT))


# ─────────────────────────────────────────────────────────────────────────────
#  3. Auto-update engine + UI files (all-or-nothing)
# ─────────────────────────────────────────────────────────────────────────────
def _bundled_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS",
                            os.path.dirname(sys.executable)))
    return Path(__file__).resolve().parent.parent  # repo root in dev


def _load_webhook() -> str | None:
    """Read the build-time error-report webhook (bundled `_webhook.cfg`,
    written by CI from the SLACK_ERROR_WEBHOOK secret — never in source).
    Expose it to the UI via ``RT_ERROR_WEBHOOK``. No-op if absent."""
    try:
        for cand in (_bundled_dir() / "_webhook.cfg",
                     Path(__file__).resolve().parent / "_webhook.cfg"):
            if cand.exists():
                url = cand.read_text().strip()
                if url:
                    os.environ["RT_ERROR_WEBHOOK"] = url
                    return url
    except Exception:
        pass
    return None


def _post_slack(text: str) -> None:
    """Fire-and-forget Slack post for LAUNCHER-side failures. Never raises."""
    url = os.environ.get("RT_ERROR_WEBHOOK")
    if not url:
        return
    try:
        import json as _json
        req = urllib.request.Request(
            url,
            data=_json.dumps({"text": text}).encode(),
            headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=4)
    except Exception:
        pass


def _fetch(url: str, timeout: int = 15) -> bytes | None:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": APP_NAME})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            if resp.status != 200:
                return None
            return resp.read()
    except (urllib.error.URLError, socket.timeout, ConnectionError):
        return None


def _validate_py(data: bytes, rel_path: str) -> bool:
    """Cheap sanity check: non-empty + has a 'def ' + compiles."""
    if not data:
        return False
    if rel_path.endswith(".py"):
        if b"def " not in data:
            return False
        try:
            compile(data, rel_path, "exec")
        except SyntaxError:
            return False
    return True


def _try_auto_update(staging: Path) -> bool:
    """Download every ENGINE_FILE into staging. True only if every file
    fetched + validated; otherwise staging is discarded and the caller falls
    back to bundled copies."""
    if staging.exists():
        shutil.rmtree(staging, ignore_errors=True)
    staging.mkdir(parents=True, exist_ok=True)
    for rel in ENGINE_FILES:
        url = RAW_URL_FMT.format(path=rel)
        data = _fetch(url)
        if data is None:
            print(f"auto-update: fetch failed for {rel}")
            return False
        if not _validate_py(data, rel):
            print(f"auto-update: validation failed for {rel}")
            return False
        target = staging / rel
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(data)
    return True


def _prepare_engine() -> tuple[Path, str]:
    bundle = _bundled_dir()
    staging = Path.home() / APP_DIR_NAME / "engine"
    print(f"auto-update: attempting fetch from {GH_OWNER}/{GH_REPO}@{GH_BRANCH}")
    if _try_auto_update(staging):
        print(f"auto-update: ok — using {staging}")
        return staging, "latest (auto-updated)"
    print(f"auto-update: falling back to bundled engine at {bundle}")
    return bundle, "bundled (offline)"


# ─────────────────────────────────────────────────────────────────────────────
#  4. Browser opener (poll /_stcore/health first)
# ─────────────────────────────────────────────────────────────────────────────
def _health_ok() -> bool:
    try:
        with urllib.request.urlopen(HEALTH_URL, timeout=2) as resp:
            return resp.status == 200 and resp.read().strip() == b"ok"
    except (urllib.error.URLError, socket.timeout, ConnectionError):
        return False


def _open_browser_when_ready() -> None:
    deadline = time.time() + 90
    while time.time() < deadline:
        if _health_ok():
            try:
                webbrowser.open(APP_URL)
            except Exception as exc:
                print(f"webbrowser.open failed: {exc}")
            return
        time.sleep(0.5)
    print("browser opener: server never returned ok within 90s")


# ─────────────────────────────────────────────────────────────────────────────
#  5. Already-running check
# ─────────────────────────────────────────────────────────────────────────────
def _already_running() -> bool:
    return _health_ok()


# ─────────────────────────────────────────────────────────────────────────────
#  6. Main — boot Streamlit
# ─────────────────────────────────────────────────────────────────────────────
def _resolve_ui_script(engine_dir: Path) -> str:
    """Prefer the freshly-downloaded ribbon_tracker_app.py; else the bundled
    copy; else this file's sibling (dev mode)."""
    for cand in (engine_dir / "ribbon_tracker_app.py",
                 _bundled_dir() / "ribbon_tracker_app.py",
                 Path(__file__).resolve().parent.parent / "ribbon_tracker_app.py"):
        if cand.exists():
            return str(cand)
    return str(engine_dir / "ribbon_tracker_app.py")


def _selftest() -> int:
    """Exercise the FULL pipeline inside the frozen process and exit 0/1.

    /_stcore/health proves only that the server booted; it never parses a
    report, renders a chart, or writes the Excel. Those run-time paths pull
    bundled data files (openpyxl write templates, altair→vega→jsonschema→
    rfc3987 grammar) that a Windows build can be missing while still serving
    health=ok. This routine runs every one of them so CI catches a packaging
    gap before techs do. Writes results to ~/.ribbonTracker/selftest.log and
    returns the process exit code (the only signal a windowed exe can give)."""
    log_dir = Path.home() / APP_DIR_NAME
    log_dir.mkdir(parents=True, exist_ok=True)
    log = open(log_dir / "selftest.log", "w", buffering=1,
               encoding="utf-8", errors="replace")
    sys.stdout = log
    sys.stderr = log

    def _ok(msg):
        print(f"[selftest] OK  {msg}")

    try:
        print(f"=== {APP_NAME} self-test {time.strftime('%Y-%m-%d %H:%M:%S')} ===")
        print(f"frozen={getattr(sys, 'frozen', False)}  python={sys.version.split()[0]}")

        # make the bundled engine importable
        sys.path.insert(0, str(_bundled_dir()))

        import io
        import tempfile
        import importlib

        import numpy as np            # noqa: F401
        import pandas as pd
        import altair as alt
        import openpyxl
        from openpyxl.styles import PatternFill
        _ok("imports: numpy / pandas / altair / openpyxl")

        ribbon_parser = importlib.import_module("ribbon_parser")
        _ok("import: ribbon_parser (bundled engine)")

        # tkinter backs the Browse folder/zip pickers — confirm it's bundled.
        # Import only (no Tk() — CI has no display); a missing module here means
        # the Browse buttons would be dead on a tech's machine.
        import tkinter            # noqa: F401
        from tkinter import filedialog  # noqa: F401
        _ok("import: tkinter + filedialog (Browse pickers)")

        # 1) openpyxl WRITE: build a synthetic report workbook
        tmp = Path(tempfile.mkdtemp(prefix="rt_selftest_"))
        rep = tmp / "report.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Acquisition Parameters"
        sr = wb.create_sheet("Splice Report")
        sr.cell(3, 1, "Ribbon"); sr.cell(3, 3, "Splice 1"); sr.cell(3, 5, "Splice 2")
        sr.cell(4, 1, "Fiber 1-12 (1) (A1)")
        sr.cell(5, 1, "Fiber 13-24 (2) (A2)")
        c = sr.cell(4, 3, "5 .312"); c.fill = PatternFill("solid", fgColor="FFC7CE")
        c = sr.cell(5, 5, "20 BEND .1 bidi"); c.fill = PatternFill("solid", fgColor="FFEB3B")
        wb.create_sheet("Legend")
        wb.save(rep)
        _ok("openpyxl write")

        # 2) parser READ + classify
        rp = ribbon_parser.parse_report(str(rep), report_name="Self Test Route")
        assert rp.grid_sheet == "Splice Report", rp.grid_sheet
        assert len(rp.events) == 2, [e.text for e in rp.events]
        _ok(f"parse_report: {len(rp.events)} events, grid={rp.grid_sheet!r}")

        # 3) pandas roll-up
        df = pd.DataFrame([e.as_row() for e in rp.events])
        g = df.groupby("ribbon_num").size()
        assert int(g.sum()) == 2
        _ok("pandas groupby roll-up")

        # 4) pandas + openpyxl WRITE path (the export)
        out_xlsx = tmp / "out.xlsx"
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xl:
            df.to_excel(xl, sheet_name="All Events", index=False)
        assert out_xlsx.exists() and out_xlsx.stat().st_size > 0
        # read it back to be sure the written file is valid
        openpyxl.load_workbook(out_xlsx)
        _ok("pandas->openpyxl ExcelWriter + reread")

        # 5) altair -> vega spec (exercises jsonschema / rfc3987 grammar)
        chart = alt.Chart(df).mark_bar().encode(
            x="ribbon_num:N", y="count():Q")
        spec = chart.to_dict()
        assert isinstance(spec, dict) and spec.get("mark")
        _ok("altair chart.to_dict() (vega/jsonschema/rfc3987)")

        import shutil
        shutil.rmtree(tmp, ignore_errors=True)
        print("=== SELF-TEST PASSED ===")
        return 0
    except BaseException as exc:  # noqa: BLE001
        import traceback
        print("=== SELF-TEST FAILED ===")
        print(f"{type(exc).__name__}: {exc}")
        print(traceback.format_exc())
        return 1


def main() -> int:
    if "--selftest" in sys.argv or os.environ.get("RT_SELFTEST"):
        return _selftest()

    _redirect_output_to_log()
    _silence_first_run_prompt()
    _load_webhook()

    if _already_running():
        print("Another instance is already serving — opening new tab.")
        try:
            webbrowser.open(APP_URL)
        except Exception:
            pass
        return 0

    engine_dir, source_label = _prepare_engine()
    os.environ["RT_ENGINE_DIR"]    = str(engine_dir)
    os.environ["RT_ENGINE_SOURCE"] = source_label

    ui_script = _resolve_ui_script(engine_dir)
    print(f"UI script: {ui_script}")

    threading.Thread(target=_open_browser_when_ready, daemon=True).start()

    from streamlit.web import cli as stcli
    sys.argv = [
        "streamlit", "run", ui_script,
        "--server.headless=true",
        f"--server.port={PORT}",
        f"--server.address={HOST}",
        "--browser.gatherUsageStats=false",
        "--global.developmentMode=false",
    ]
    try:
        return stcli.main()
    except SystemExit as exc:
        return exc.code if isinstance(exc.code, int) else 0
    except BaseException as exc:
        import platform
        import traceback
        try:
            _post_slack(
                f":rotating_light: *Ribbon Tracker launcher* failed to start\n"
                f"*{type(exc).__name__}*: {exc}\n"
                f"host: `{platform.node()}` | os: {platform.platform()} | "
                f"engine: {os.environ.get('RT_ENGINE_SOURCE','?')}\n"
                f"```{traceback.format_exc()[-1400:]}```")
        except Exception:
            pass
        raise


if __name__ == "__main__":
    sys.exit(main() or 0)
