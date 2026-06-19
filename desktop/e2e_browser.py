"""
e2e_browser.py — real browser click-through against a RUNNING Ribbon Tracker.

Drives a headless Chromium through the exact flow a tech (Zach) performs:
  1. open the app the frozen exe is serving,
  2. UPLOAD two reports through the real file uploader,
  3. wait for the metrics + ribbon ranking + repeat-offender callout to render,
  4. confirm a timestamped Excel was auto-saved to disk,
  5. screenshot the result.

The app must already be running on BASE_URL (the CI workflow launches the .exe
first). Exits non-zero with a message on any failure. This is the gate that
proves the whole upload->analyze->chart->save round-trip works in the browser
served by the FROZEN build, not just that the libraries import.

Usage: python e2e_browser.py [base_url] [screenshot_path]
"""
from __future__ import annotations

import os
import re
import sys
import tempfile

import openpyxl
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright, expect

BASE_URL = sys.argv[1] if len(sys.argv) > 1 else "http://127.0.0.1:8512"
SHOT = sys.argv[2] if len(sys.argv) > 2 else "e2e_browser.png"
PINK = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB3B")


def make_report(path, cells):
    wb = openpyxl.Workbook()
    wb.active.title = "Acquisition Parameters"
    sr = wb.create_sheet("Splice Report")
    sr.cell(3, 1, "Ribbon"); sr.cell(3, 3, "Splice 1"); sr.cell(3, 5, "Splice 2")
    for r, lab in [(4, "Fiber 1-12 (1) (A1)"), (5, "Fiber 13-24 (2) (A2)"),
                   (6, "Fiber 25-36 (3) (B1)")]:
        sr.cell(r, 1, lab)
    for r, c, v, f in cells:
        cell = sr.cell(r, c, v)
        cell.fill = f
    wb.create_sheet("Legend")
    wb.save(path)


def main() -> int:
    tmp = tempfile.mkdtemp(prefix="rt_e2e_")
    # Ribbon 1 flagged in BOTH -> must surface as a repeat offender.
    f1 = os.path.join(tmp, "Route Alpha.xlsx")
    f2 = os.path.join(tmp, "Route Beta.xlsx")
    make_report(f1, [(4, 3, "5 .312", PINK), (6, 5, "30 BEND .1 bidi", YELLOW)])
    make_report(f2, [(4, 5, "7 .288", PINK)])

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.set_default_timeout(60000)
        print(f"[e2e] opening {BASE_URL}")
        page.goto(BASE_URL, wait_until="domcontentloaded")
        # app header
        expect(page.get_by_text("Ribbon Issue Tracker").first).to_be_visible()

        # UPLOAD through the real uploader (hidden <input type=file>, multiple)
        print("[e2e] uploading 2 reports")
        page.locator('input[type="file"]').set_input_files([f1, f2])

        # metrics + repeat-offender callout must render
        expect(page.get_by_text("Reports loaded").first).to_be_visible()
        expect(page.get_by_text("Total flagged events").first).to_be_visible()
        expect(page.get_by_text("Repeat offender").first).to_be_visible(timeout=60000)
        expect(page.get_by_text("Ribbon ranking").first).to_be_visible()
        # the export section is at the very bottom — wait for it to render
        expect(page.get_by_text("Report saved").first).to_be_visible(timeout=60000)

        body = page.inner_text("body")
        page.screenshot(path=SHOT, full_page=True)
        browser.close()

    # sanity on what rendered
    assert "Reports loaded" in body, "metrics missing"
    assert "Repeat offender" in body and "Ribbon 1" in body, "repeat-offender wrong"

    # the UI prints the saved path; verify that exact file exists + is valid
    m = re.search(r"Report saved:\s*`?(.+?\.xlsx)`?", body, re.S)
    if not m:
        print("[e2e] FAIL — no 'Report saved' path on the page")
        return 1
    saved = m.group(1).strip().strip("`").strip()
    if not os.path.exists(saved):
        print(f"[e2e] FAIL — saved report not found on disk: {saved}")
        return 1
    wb = openpyxl.load_workbook(saved)
    sheets = wb.sheetnames
    # clean up so the test never leaves a file in the user's reports folder
    try:
        os.remove(saved)
    except OSError:
        pass
    if sheets != ["Ribbon Ranking", "All Events", "Reports Loaded"]:
        print(f"[e2e] FAIL — unexpected sheets: {sheets}")
        return 1
    print(f"[e2e] OK — report auto-saved + valid: {os.path.basename(saved)}")
    print(f"[e2e] screenshot: {SHOT}")
    print("[e2e] PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
